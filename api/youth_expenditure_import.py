"""Read-only import helpers for Youth Jobs expenditure actuals.

The management workbook remains the source of truth. This module selects a dated
workbook, verifies that it stays byte-identical while being read, and returns a
small monthly snapshot for an explicit database publication step.
"""

from __future__ import annotations

import calendar
import hashlib
import re
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook


MONTHS = {
    name.lower(): number
    for number, name in enumerate(calendar.month_name)
    if name
}
MANAGEMENT_WORKBOOK_RE = re.compile(
    r"^(?P<date>\d{8}) - .*Management Accounts.*\.xlsx$",
    re.IGNORECASE,
)
EXCEL_ERROR_VALUES = {
    "#DIV/0!",
    "#N/A",
    "#NAME?",
    "#NULL!",
    "#NUM!",
    "#REF!",
    "#VALUE!",
}
REQUIRED_HEADERS = (
    "Date",
    "Month",
    "Year",
    "Amount",
    "Category 1",
    "Category 2",
    "Category 3",
)


class YouthExpenditureImportError(ValueError):
    """The workbook cannot safely produce a publishable monthly snapshot."""


@dataclass(frozen=True)
class MonthlyYouthExpenditureTotal:
    month: int
    core_amount: Decimal
    mentor_amount: Decimal
    rural_amount: Decimal
    row_count: int

    @property
    def total(self) -> Decimal:
        return self.core_amount + self.mentor_amount + self.rural_amount


@dataclass(frozen=True)
class CategoryError:
    row: int
    month: int | None
    amount: Decimal
    values: tuple[str, str, str]


@dataclass(frozen=True)
class YouthExpenditureSnapshot:
    path: Path
    year: int
    sha256: str
    sha256_after_read: str
    source_size: int
    source_modified_at: datetime
    months: dict[int, MonthlyYouthExpenditureTotal]
    classified_row_count: int
    category_errors: tuple[CategoryError, ...]
    date_period_mismatch_rows: tuple[int, ...]


def resolve_latest_management_workbook(directory: Path | str) -> Path:
    """Return the newest YYYYMMDD management workbook from ``directory``.

    The date prefix is authoritative. If several copies share the newest date,
    the most recently modified copy wins, with the filename as a deterministic
    final tie-breaker.
    """

    directory = Path(directory).expanduser().resolve()
    if not directory.is_dir():
        raise YouthExpenditureImportError(
            f"Management workbook directory does not exist: {directory}"
        )

    candidates: list[tuple[str, Path]] = []
    for path in directory.iterdir():
        match = MANAGEMENT_WORKBOOK_RE.match(path.name)
        if path.is_file() and match and not path.name.startswith("~$"):
            candidates.append((match.group("date"), path))
    if not candidates:
        raise YouthExpenditureImportError(
            f"No dated management workbook found in {directory}"
        )

    newest_date = max(date_prefix for date_prefix, _ in candidates)
    newest_candidates = [
        path for date_prefix, path in candidates if date_prefix == newest_date
    ]
    return max(
        newest_candidates,
        key=lambda path: (path.stat().st_mtime_ns, path.name),
    )


def parse_amount(raw) -> Decimal:
    """Parse a finance-ledger amount without introducing binary rounding."""

    cleaned = (
        str(raw or "")
        .strip()
        .replace("R", "")
        .replace(",", "")
        .replace(" ", "")
    )
    if not cleaned or cleaned == "-":
        return Decimal("0")
    try:
        return Decimal(cleaned)
    except InvalidOperation as exc:
        raise YouthExpenditureImportError(
            f"Invalid Amount value: {raw!r}"
        ) from exc


def bucket_for(row: dict[str, object]) -> str:
    """Classify a ledger payment using the canonical priority order."""

    category_2 = str(row.get("Category 2") or "").casefold()
    category_3 = str(row.get("Category 3") or "").casefold()
    if "mentor" in category_3:
        return "mentor_amount"
    if "wind farm" in category_2 or "rural" in category_3:
        return "rural_amount"
    return "core_amount"


def read_youth_expenditure_workbook(
    path: Path | str,
    *,
    year: int,
) -> YouthExpenditureSnapshot:
    """Aggregate one year's classified Youth Jobs expenditure by month."""

    path = Path(path).expanduser().resolve()
    if not path.is_file():
        raise YouthExpenditureImportError(f"Workbook does not exist: {path}")

    stat_before = path.stat()
    sha_before = _sha256(path)
    workbook = load_workbook(
        path,
        read_only=True,
        data_only=True,
        keep_links=True,
    )
    try:
        if "Expenditure" not in workbook.sheetnames:
            raise YouthExpenditureImportError(
                f"Workbook has no Expenditure sheet: {path.name}"
            )
        sheet = workbook["Expenditure"]
        header_cells = next(
            sheet.iter_rows(min_row=1, max_row=1, min_col=1, max_col=12)
        )
        headers = {
            str(cell.value).strip(): index
            for index, cell in enumerate(header_cells)
            if cell.value not in (None, "")
        }
        missing_headers = [
            header for header in REQUIRED_HEADERS if header not in headers
        ]
        if missing_headers:
            raise YouthExpenditureImportError(
                "Expenditure sheet is missing required headers: "
                + ", ".join(missing_headers)
            )

        accumulators: dict[int, dict[str, Decimal | int]] = {}
        category_errors: list[CategoryError] = []
        date_period_mismatch_rows: list[int] = []

        for row_number, cells in enumerate(
            sheet.iter_rows(min_row=2, min_col=1, max_col=12),
            start=2,
        ):
            values = [cell.value for cell in cells]
            if all(value in (None, "") for value in values):
                continue

            row_year = _integer(values[headers["Year"]])
            if row_year != year:
                continue

            raw_month = values[headers["Month"]]
            month = _month_number(raw_month)
            category_cells = tuple(
                cells[headers[header]]
                for header in ("Category 1", "Category 2", "Category 3")
            )
            if any(_is_excel_error(cell.value, cell.data_type) for cell in category_cells):
                amount = parse_amount(values[headers["Amount"]])
                category_errors.append(
                    CategoryError(
                        row=row_number,
                        month=month,
                        amount=amount,
                        values=tuple(str(cell.value or "") for cell in category_cells),
                    )
                )
                continue

            category_1, category_2, category_3 = (
                str(cell.value or "").strip() for cell in category_cells
            )
            if "youth jobs:" not in category_3.casefold():
                continue
            amount = parse_amount(values[headers["Amount"]])
            if category_1.casefold() != "children & youth":
                raise YouthExpenditureImportError(
                    f"Youth Jobs row {row_number} has unexpected Category 1: "
                    f"{category_1!r}"
                )
            if month is None:
                raise YouthExpenditureImportError(
                    f"Youth Jobs row {row_number} has invalid Month: {raw_month!r}"
                )

            row_date = values[headers["Date"]]
            if isinstance(row_date, datetime):
                row_date = row_date.date()
            if isinstance(row_date, date) and (
                row_date.year != year or row_date.month != month
            ):
                date_period_mismatch_rows.append(row_number)

            accumulator = accumulators.setdefault(
                month,
                {
                    "core_amount": Decimal("0"),
                    "mentor_amount": Decimal("0"),
                    "rural_amount": Decimal("0"),
                    "row_count": 0,
                },
            )
            bucket = bucket_for(
                {
                    "Category 2": category_2,
                    "Category 3": category_3,
                }
            )
            accumulator[bucket] += amount
            accumulator["row_count"] += 1
    finally:
        workbook.close()

    stat_after = path.stat()
    sha_after = _sha256(path)
    if (
        sha_before != sha_after
        or stat_before.st_size != stat_after.st_size
        or stat_before.st_mtime_ns != stat_after.st_mtime_ns
    ):
        raise YouthExpenditureImportError(
            f"Workbook changed during import: {path}"
        )
    if not accumulators:
        raise YouthExpenditureImportError(
            f"No classified Youth Jobs rows found for {year} in {path.name}"
        )

    actual_months = set(accumulators)
    expected_months = set(range(1, max(actual_months) + 1))
    if actual_months != expected_months:
        missing = ", ".join(
            calendar.month_name[month]
            for month in sorted(expected_months - actual_months)
        )
        raise YouthExpenditureImportError(
            f"Actual months for {year} are not contiguous from January; "
            f"missing: {missing}"
        )

    months = {
        month: MonthlyYouthExpenditureTotal(
            month=month,
            core_amount=accumulator["core_amount"],
            mentor_amount=accumulator["mentor_amount"],
            rural_amount=accumulator["rural_amount"],
            row_count=accumulator["row_count"],
        )
        for month, accumulator in sorted(accumulators.items())
    }
    return YouthExpenditureSnapshot(
        path=path,
        year=year,
        sha256=sha_before,
        sha256_after_read=sha_after,
        source_size=stat_before.st_size,
        source_modified_at=datetime.fromtimestamp(stat_before.st_mtime),
        months=months,
        classified_row_count=sum(month.row_count for month in months.values()),
        category_errors=tuple(category_errors),
        date_period_mismatch_rows=tuple(date_period_mismatch_rows),
    )


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _integer(raw) -> int | None:
    try:
        return int(float(str(raw).strip()))
    except (TypeError, ValueError):
        return None


def _month_number(raw) -> int | None:
    if isinstance(raw, bool) or raw is None:
        return None
    if isinstance(raw, (int, float)):
        month = int(raw)
        return month if 1 <= month <= 12 else None
    return MONTHS.get(str(raw).strip().casefold())


def _is_excel_error(value, data_type: str) -> bool:
    return data_type == "e" or str(value or "").strip().upper() in EXCEL_ERROR_VALUES
