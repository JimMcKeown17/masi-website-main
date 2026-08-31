from __future__ import annotations

import os
from datetime import date
from decimal import Decimal
from io import StringIO
from pathlib import Path
from tempfile import TemporaryDirectory

from django.core.management import call_command
from django.core.management.base import CommandError
from django.test import TestCase
from openpyxl import Workbook

from api.models import MonthlyYouthExpenditure
from api.youth_expenditure_import import (
    YouthExpenditureImportError,
    read_youth_expenditure_workbook,
    resolve_latest_management_workbook,
)


HEADERS = (
    "Receipt #",
    "Date",
    "Month",
    "Year",
    "FY",
    "ID",
    "Name",
    "Amount",
    "Paid By",
    "Category 1",
    "Category 2",
    "Category 3",
)


def youth_row(
    *,
    receipt: str,
    row_date: date,
    month: str,
    year: int,
    amount: Decimal,
    category_2: str = "Literacy Projects",
    category_3: str = "Youth Jobs: Primary Literacy Coach",
):
    return (
        receipt,
        row_date,
        month,
        year,
        f"FY{year + 1}",
        "1001",
        "Fixture youth",
        amount,
        "FNB",
        "Children & Youth",
        category_2,
        category_3,
    )


def write_workbook(
    path: Path,
    *,
    include_category_error: bool = False,
    include_month_gap: bool = False,
):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Expenditure"
    sheet.append(HEADERS)
    sheet.append(
        youth_row(
            receipt="A1",
            row_date=date(2026, 1, 30),
            month="January",
            year=2026,
            amount=Decimal("100.00"),
        )
    )
    sheet.append(
        youth_row(
            receipt="A2",
            row_date=date(2026, 1, 30),
            month="January",
            year=2026,
            amount=Decimal("200.00"),
            category_3="Youth Jobs: LC Mentor",
        )
    )
    sheet.append(
        youth_row(
            receipt="A3",
            row_date=date(2026, 1, 30),
            month="January",
            year=2026,
            amount=Decimal("300.00"),
            category_2="Kouga Wind Farm",
        )
    )
    sheet.append(
        youth_row(
            receipt="A4",
            row_date=date(2026, 1, 30),
            month="January",
            year=2026,
            amount=Decimal("400.00"),
            category_2="Kouga Wind Farm",
            category_3="Youth Jobs: LC Mentor (Rural)",
        )
    )
    second_month = "March" if include_month_gap else "February"
    second_month_number = 3 if include_month_gap else 2
    sheet.append(
        youth_row(
            receipt="B1",
            row_date=date(2026, second_month_number, 27),
            month=second_month,
            year=2026,
            amount=Decimal("50.25"),
        )
    )
    sheet.append(
        youth_row(
            receipt="OLD",
            row_date=date(2025, 12, 30),
            month="December",
            year=2025,
            amount=Decimal("999.00"),
        )
    )
    sheet.append(
        (
            "OTHER",
            date(2026, 2, 20),
            "February",
            2026,
            "FY2027",
            "",
            "Non-youth fixture",
            "not an amount",
            "FNB",
            "Operations",
            "Office",
            "Stationery",
        )
    )
    if include_category_error:
        sheet.append(
            (
                "ERR",
                date(2026, 2, 25),
                "February",
                2026,
                "FY2027",
                "",
                "Unclassified fixture",
                Decimal("700.00"),
                "FNB",
                "#NAME?",
                "#NAME?",
                "#NAME?",
            )
        )
        for column in (10, 11, 12):
            sheet.cell(row=sheet.max_row, column=column).data_type = "e"
    workbook.save(path)
    workbook.close()


class WorkbookSelectionTests(TestCase):
    def test_selects_latest_date_then_newest_copy_within_that_date(self):
        with TemporaryDirectory() as temporary_directory:
            directory = Path(temporary_directory)
            older = directory / "20260828 - Masinyusane Management Accounts.xlsx"
            first_latest = directory / "20260829 - Masinyusane Management Accounts A.xlsx"
            newest_latest = directory / "20260829 - Masinyusane Management Accounts B.xlsx"
            ignored = directory / "notes.xlsx"
            for path in (older, first_latest, newest_latest, ignored):
                path.touch()
            os.utime(first_latest, ns=(1_000_000_000, 1_000_000_000))
            os.utime(newest_latest, ns=(2_000_000_000, 2_000_000_000))

            selected = resolve_latest_management_workbook(directory)

            self.assertEqual(selected, newest_latest.resolve())

    def test_rejects_directory_without_a_dated_management_workbook(self):
        with TemporaryDirectory() as temporary_directory:
            with self.assertRaisesRegex(
                YouthExpenditureImportError,
                "No dated management workbook",
            ):
                resolve_latest_management_workbook(Path(temporary_directory))


class WorkbookAggregationTests(TestCase):
    def test_aggregates_youth_jobs_by_month_with_mentor_priority(self):
        with TemporaryDirectory() as temporary_directory:
            path = (
                Path(temporary_directory)
                / "20260829 - Masinyusane Management Accounts.xlsx"
            )
            write_workbook(path, include_category_error=True)

            snapshot = read_youth_expenditure_workbook(path, year=2026)

            january = snapshot.months[1]
            self.assertEqual(january.core_amount, Decimal("100.00"))
            self.assertEqual(january.mentor_amount, Decimal("600.00"))
            self.assertEqual(january.rural_amount, Decimal("300.00"))
            self.assertEqual(january.total, Decimal("1000.00"))
            self.assertEqual(january.row_count, 4)
            self.assertEqual(snapshot.months[2].total, Decimal("50.25"))
            self.assertEqual(snapshot.classified_row_count, 5)
            self.assertEqual(len(snapshot.category_errors), 1)
            self.assertEqual(snapshot.category_errors[0].amount, Decimal("700.00"))
            self.assertEqual(snapshot.sha256, snapshot.sha256_after_read)

    def test_rejects_a_gap_in_the_actual_month_series(self):
        with TemporaryDirectory() as temporary_directory:
            path = (
                Path(temporary_directory)
                / "20260829 - Masinyusane Management Accounts.xlsx"
            )
            write_workbook(path, include_month_gap=True)

            with self.assertRaisesRegex(
                YouthExpenditureImportError,
                "contiguous",
            ):
                read_youth_expenditure_workbook(path, year=2026)


class SyncYouthExpenditureCommandTests(TestCase):
    def setUp(self):
        self.temporary_directory = TemporaryDirectory()
        self.directory = Path(self.temporary_directory.name)
        self.path = (
            self.directory / "20260829 - Masinyusane Management Accounts.xlsx"
        )
        write_workbook(self.path, include_category_error=True)
        MonthlyYouthExpenditure.objects.create(
            year=2026,
            month=1,
            core_amount=Decimal("999.00"),
            note="old snapshot",
        )

    def tearDown(self):
        self.temporary_directory.cleanup()

    def test_dry_run_reports_differences_without_writing(self):
        output = StringIO()

        call_command(
            "sync_youth_expenditure",
            workbook_dir=str(self.directory),
            year=2026,
            stdout=output,
        )

        january = MonthlyYouthExpenditure.objects.get(year=2026, month=1)
        self.assertEqual(january.core_amount, Decimal("999.00"))
        self.assertFalse(
            MonthlyYouthExpenditure.objects.filter(year=2026, month=2).exists()
        )
        rendered = output.getvalue()
        self.assertIn("DRY RUN", rendered)
        self.assertIn(self.path.name, rendered)
        self.assertIn("Category errors: 1", rendered)

    def test_apply_requires_explicit_category_error_acknowledgement(self):
        with self.assertRaisesRegex(CommandError, "category errors"):
            call_command(
                "sync_youth_expenditure",
                workbook_dir=str(self.directory),
                year=2026,
                apply=True,
                stdout=StringIO(),
                stderr=StringIO(),
            )

        january = MonthlyYouthExpenditure.objects.get(year=2026, month=1)
        self.assertEqual(january.core_amount, Decimal("999.00"))

    def test_apply_restates_existing_months_and_records_provenance(self):
        call_command(
            "sync_youth_expenditure",
            workbook_dir=str(self.directory),
            year=2026,
            apply=True,
            allow_category_errors=True,
            stdout=StringIO(),
        )

        january = MonthlyYouthExpenditure.objects.get(year=2026, month=1)
        february = MonthlyYouthExpenditure.objects.get(year=2026, month=2)
        self.assertEqual(january.core_amount, Decimal("100.00"))
        self.assertEqual(january.mentor_amount, Decimal("600.00"))
        self.assertEqual(january.rural_amount, Decimal("300.00"))
        self.assertEqual(february.core_amount, Decimal("50.25"))
        self.assertIn(self.path.name, january.note)
        self.assertIn("sha256=", january.note)
        self.assertIn("rows=4", january.note)

        call_command(
            "sync_youth_expenditure",
            workbook_dir=str(self.directory),
            year=2026,
            apply=True,
            allow_category_errors=True,
            stdout=StringIO(),
        )

        self.assertEqual(
            MonthlyYouthExpenditure.objects.filter(year=2026).count(),
            2,
        )
